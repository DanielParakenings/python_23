import logging
import os
import requests
import json
import azure.functions as func
from azure.identity import DefaultAzureCredential
from azure.keyvault.secrets import SecretClient
from openpyxl import Workbook
from azure.storage.blob import BlobServiceClient
from datetime import datetime

def get_secret(key_vault_uri, secret_name):
    credential = DefaultAzureCredential()
    client = SecretClient(vault_url=key_vault_uri, credential=credential)
    secret = client.get_secret(secret_name)
    return secret.value

def acquire_token(tenant_id, client_id, client_secret):
    token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
    token_data = {
        'grant_type': 'client_credentials',
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'https://analysis.windows.net/powerbi/api/.default'
    }
    token_r = requests.post(token_url, data=token_data)
    return token_r.json().get("access_token")

def query_powerbi(token, group_id, dataset_id, dax_query):
    url = f"https://api.powerbi.com/v1.0/myorg/groups/{group_id}/datasets/{dataset_id}/executeQueries"
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json'
    }
    response = requests.post(url, headers=headers, json=dax_query)
    response.raise_for_status()
    return response.json()

def main(mytimer: func.TimerRequest) -> None:
    logging.info('Python timer trigger function started.')

    key_vault_uri = os.environ["KEY_VAULT_URI"]

    # Abrufen der Anmeldeinformationen aus dem Key Vault
    client_id = get_secret(key_vault_uri, "CLIENT_ID")
    tenant_id = get_secret(key_vault_uri, "TENANT_ID")
    client_secret = get_secret(key_vault_uri, "CLIENT_SECRET")
    storage_account_name = get_secret(key_vault_uri, "STORAGE_ACCOUNT_NAME")
    storage_account_key = get_secret(key_vault_uri, "STORAGE_ACCOUNT_KEY")

    # Abrufen des Tokens
    token = acquire_token(tenant_id, client_id, client_secret)

    group_id = "YOUR_GROUP_ID"  # Arbeitsbereichs-ID
    dataset_id = "YOUR_DATASET_ID"  # Dataset-ID

    # Erste DAX-Abfrage für distinct Account Names
    dax_query_accounts = {
        "queries": [
            {
                "query": "EVALUATE DISTINCT('Table'[Account_Name])",
                "name": "AccountNamesQuery"
            }
        ]
    }

    # Ausführen der ersten DAX-Abfrage
    account_names_result = query_powerbi(token, group_id, dataset_id, dax_query_accounts)
    account_names = [row['Account_Name'] for row in account_names_result['results'][0]['tables'][0]['rows']]

    # Excel-Arbeitsmappe erstellen
    workbook = Workbook()
    for account_name in account_names:
        # Blatt für jeden Account Name erstellen
        sheet = workbook.create_sheet(title=account_name)

        # Zweite DAX-Abfrage für Details zu jedem Account Name
        dax_query_details = {
            "queries": [
                {
                    "query": f"EVALUATE FILTER('Table', 'Table'[Account_Name] = \"{account_name}\")",
                    "name": f"DetailsQuery_{account_name}"
                }
            ]
        }

        # Ausführen der zweiten DAX-Abfrage
        details_result = query_powerbi(token, group_id, dataset_id, dax_query_details)
        rows = details_result['results'][0]['tables'][0]['rows']

        # Schreiben der Daten in das Excel-Blatt
        if rows:
            headers = rows[0].keys()
            sheet.append(headers)
            for row in rows:
                sheet.append(row.values())

    # Entfernen des Standardblatts
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    # Speichern der Excel-Datei
    file_name = f"PowerBI_Data_{datetime.now().strftime('%Y%m%d')}.xlsx"
    workbook.save(file_name)

    # Hochladen der Excel-Datei in Azure Blob Storage
    blob_service_client = BlobServiceClient(
        account_url=f"https://{storage_account_name}.blob.core.windows.net",
        credential=storage_account_key
    )
    container_name = "your-container-name"
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=file_name)

    with open(file_name, "rb") as data:
        blob_client.upload_blob(data, overwrite=True)

    logging.info('Python timer trigger function completed successfully.')
